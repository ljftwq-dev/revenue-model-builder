"""Excel builder — render a RevenueModel into a formatted .xlsx.

Implements the ABC color coding, IF-protected revenue formulas, and the
residual line. Historical and forecast years are physically separated:
forecast columns get an orange tint and are left blank (structure ready,
values filled once the historical model ties out — the "history first,
then forecast" workflow).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .driver import LEVEL_A, LEVEL_B, LEVEL_C, PENETRATION, SHARE, BASE, PRICE
from .model import RevenueModel
from .segment import Segment

COLOR_A = "000000"
COLOR_B = "0000FF"
COLOR_C = "FF0000"

_LEVEL_COLOR = {LEVEL_A: COLOR_A, LEVEL_B: COLOR_B, LEVEL_C: COLOR_C}

_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FILL_FORECAST = PatternFill("solid", fgColor="FFF3E0")
ALIGN_C = Alignment(horizontal="center")


def _num_format(driver):
    if driver.kind in (PENETRATION, SHARE):
        return "0.00%"
    if driver.kind == BASE:
        return "#,##0.00"
    if driver.kind == PRICE:
        return "#,##0"
    return "#,##0.00"


def build_excel(model: RevenueModel, path: str, forecast_years=None):
    hist = model.years()
    fcast = list(forecast_years or [])
    years = hist + fcast
    Y0 = 4

    def ycol(y):
        return Y0 + years.index(y)

    def is_forecast(y):
        return y in fcast

    wb = Workbook()
    ws = wb.active
    ws.title = (model.company or "model")[:31]

    ws.cell(1, 1, f"{model.company} 收入模型").font = Font(
        bold=True, size=14, name="微软雅黑", color="1F3A5F")

    for ci, label in enumerate(["项目", "单位", "等级"], start=1):
        ws.cell(2, ci, label).font = Font(bold=True)
    for y in years:
        cell = ws.cell(2, ycol(y), f"{y}E" if is_forecast(y) else str(y))
        cell.font = Font(bold=True)
        cell.alignment = ALIGN_C
        if is_forecast(y):
            cell.fill = FILL_FORECAST
        cell.border = BORDER

    row = 3
    revenue_rows = []

    for seg in model.segments:
        ws.cell(row, 1, seg.name).font = Font(bold=True, size=11, color="2E5C8A")
        row += 1
        driver_start = row
        for d in seg.drivers():
            color = _LEVEL_COLOR.get(d.level, COLOR_C)
            ws.cell(row, 1, f"  {d.kind_label()}").font = Font(color=color)
            ws.cell(row, 2, d.unit).font = Font(color=color)
            ws.cell(row, 3, d.level).font = Font(color=color, bold=True)
            for y in years:
                cell = ws.cell(row, ycol(y))
                cell.number_format = _num_format(d)
                cell.border = BORDER
                cell.font = Font(color=color)
                if y in hist and y in d.values:
                    cell.value = d.values[y]
                if is_forecast(y):
                    cell.fill = FILL_FORECAST
            row += 1
        ws.cell(row, 1, "  收入").font = Font(bold=True)
        for y in years:
            col = get_column_letter(ycol(y))
            f = (f'=IF(OR({col}{driver_start}="",{col}{driver_start+1}=""),"",'
                 f'{col}{driver_start}*{col}{driver_start+1}*'
                 f'{col}{driver_start+2}*{col}{driver_start+3})')
            cell = ws.cell(row, ycol(y), f)
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=True)
            cell.border = BORDER
            if is_forecast(y):
                cell.fill = FILL_FORECAST
        revenue_rows.append(row)
        row += 2

    ws.cell(row, 1, "汇总").font = Font(bold=True, size=11, color="2E5C8A")
    row += 1

    ws.cell(row, 1, "Σ 分项").font = Font(bold=True)
    for y in years:
        col = get_column_letter(ycol(y))
        refs = ",".join(f"{col}{r}" for r in revenue_rows)
        cell = ws.cell(row, ycol(y), f"=SUM({refs})")
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True)
        cell.border = BORDER
    sum_row = row
    row += 1

    ws.cell(row, 1, "总收入（年报）").font = Font(bold=True)
    for y in years:
        cell = ws.cell(row, ycol(y))
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True, color=COLOR_A)
        cell.border = BORDER
        if y in hist:
            cell.value = model.total_revenue[y]
        if is_forecast(y):
            cell.fill = FILL_FORECAST
    total_row = row
    row += 1

    ws.cell(row, 1, "差额行").font = Font(bold=True, color="808080")
    for y in years:
        col = get_column_letter(ycol(y))
        cell = ws.cell(row, ycol(y))
        cell.number_format = "#,##0.00"
        cell.border = BORDER
        cell.font = Font(color="808080")
        if y in hist:
            cell.value = f"={col}{total_row}-{col}{sum_row}"
        if is_forecast(y):
            cell.fill = FILL_FORECAST
    row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 6
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(Y0 + i)].width = 12

    wb.save(path)
    return path


def main(argv=None):
    import argparse
    import os
    from .demo import build_novatech
    parser = argparse.ArgumentParser(
        description="Build the NovaTech fictional demo revenue model to .xlsx")
    parser.add_argument(
        "output", nargs="?", default="NovaTech_revenue_model_demo.xlsx",
        help="output .xlsx path (default: ./NovaTech_revenue_model_demo.xlsx)")
    args = parser.parse_args(argv)
    model = build_novatech()
    out = os.path.abspath(args.output)
    build_excel(model, out, forecast_years=[2025, 2026, 2027])
    print(f"OK -> {out}")
    print(f"size: {os.path.getsize(out)} bytes")
    print("historical columns: filled with real data + formulas")
    print("forecast columns 2025E/2026E/2027E: structure reserved (orange), values blank")


if __name__ == "__main__":
    main()
